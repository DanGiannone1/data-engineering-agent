import pandas as pd
import openpyxl

# Load the mapping spreadsheet
mapping_file = r"input_data\input\Anonymized\ZZzzzzzz_aaab3_Master_Mapping_Repository.xlsx"

# Read all sheet names
wb = openpyxl.load_workbook(mapping_file, read_only=True, data_only=True)
print("Sheet names in mapping file:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")

print("\n" + "="*80 + "\n")

# Try to read the first few sheets
for sheet_name in wb.sheetnames[:3]:  # Only first 3 sheets
    print(f"\nSheet: {sheet_name}")
    print("-" * 40)
    df = pd.read_excel(mapping_file, sheet_name=sheet_name, nrows=10)
    print(f"Shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    print(f"\nFirst 5 rows:")
    print(df.head())
    print("\n")
